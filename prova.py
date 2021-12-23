import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from git import Repo
import git
import logging
import traceback
import pandas as pd


file_excel = openpyxl.load_workbook('data3.xlsx')
sheet = file_excel.get_sheet_by_name('Sheet1')
repo = git.Repo('/Users/patriziadecristofaro/Desktop/Arduino')
data = []
for x in range(1,681):
    try:
        comm1=(sheet.cell(x,1).value)
        comm2=(sheet.cell(x,2).value)
        logging.info("processing "+ comm1 + " and " + comm2)
        d = repo.git.diff("--stat", comm1, comm2)
        s = d.split("\n")
        s.pop()
        for l in s:
            ss = l.split('|')
            fname = "/Users/patriziadecristofaro/Desktop/Arduino/" + ss[0].strip()
            modLines = ss[1].strip().replace("+", "").replace("-", "")
            data1 = [comm1, fname, modLines]
            data.append(data1)
    except(Exception) as e:
        logging.info(e)
        logging.info(traceback.format_exc())
df = pd.DataFrame(data, columns = ['hash', 'file', 'modlines'])
df.to_csv("DiffDataset.csv", index = False)