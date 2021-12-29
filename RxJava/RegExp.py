###
### Step 1 Progetto Evoluzione del Software
###
### Filtraggio Commit in base alle RegExp espresse 
### Filtraggio duplicati
### Creazione dataset con commit Safe e Defect
###

import re
import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from github import Github
import xlsxwriter

# Commit Con Fix (stato safe)
workbook1 = xlsxwriter.Workbook('Safe.xlsx')
worksheet1 = workbook1.add_worksheet()

# Commit Precedenti (stato defect)
workbook = xlsxwriter.Workbook('Defect.xlsx')
worksheet = workbook.add_worksheet()

# Merge safe e defect
w = xlsxwriter.Workbook('SafeDefectMerge.xlsx')
wk = w.add_worksheet()

# File dei Commit
file_excel = openpyxl.load_workbook('commit.xlsx')
sheet = file_excel.get_sheet_by_name('commit')

y=0
z=0
c=0
f=0
g=0
k=0
safeDict=dict()
defectDict=dict()


for x in range(1,3196):
    safeA=(sheet.cell(x,4).value)
    safeB=(sheet.cell(x,1).value)
    giorno=str((sheet.cell(x,3).value))

    if(x > 1): 
        defectA=(sheet.cell(x,4).value)
        defectB=(sheet.cell((x-1),1).value)

    try:
        s1=re.search(r"refact", safeA, re.I)
        s2=re.search(r"Fix", safeA, re.I)
        s3=re.search(r"bug", safeA, re.I)
        s4=re.search(r"issue", safeA, re.I)
        s5=re.search(r"remov", safeA, re.I)
        s6=re.search(r"resolv", safeA, re.I)
        
        
        if(s1 is not None):
            print(safeB)
            y=y+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno
        if(s2 is not None):
            print(safeB)
            z=z+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno
        if(s3 is not None):
            print(safeB)
            c=c+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno
        if( s4 is not None):
            print(safeB)
            f=f+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno
        if( s5 is not None):
            print(safeB)
            g=g+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno
        if( s6 is not None):
            print(safeB)
            k=k+1
            safeDict[str(safeB)]=giorno
            if(x > 1): defectDict[str(defectB)]=giorno

    except AttributeError:
        pass
        
print("regular expression with refactor = ", y)
print("regular expression with Fix = ", z)
print("regular expression with bug = ", c)
print("regular expression with issue = ", f)
print("regular expression with remov = ", g)
print("regular expression with resolv = ", k)


row=0
col=0        
wk.write(0, 0, "Defect")
wk.write(0, 1, "Safe")
for key in defectDict.keys():
    worksheet.write(row, col, key)
    worksheet.write(row,col+1, defectDict[key])
    wk.write(row+1, col, key)
    row += 1
row = 0
for key in safeDict.keys():
    worksheet1.write(row, col, key)
    wk.write(row+1, col+1, key)
    row += 1


w.close()
workbook.close()
workbook1.close()

print("stati safe totali = ", len(safeDict))
print("stati defect totali = ", len(defectDict))