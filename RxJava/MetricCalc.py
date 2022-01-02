###
### Step 2 Progetto Evoluzione del Software
###
### Calcolo Metriche per ogni commit nello stato DEFECT
### Salvataggio Metriche nei file
### Creazione dataset con Git Diff calcolo righe modificate 
### per passare dallo stato DEFECT allo stato SAFE (per fixare il bug)
###


import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from git import Repo
import git
import subprocess
import shutil
import logging
import traceback
import pandas as pd

### CALCOLO METRICHE 

file_excel = openpyxl.load_workbook('SafeDefectMerge.xlsx')
sheet = file_excel.get_sheet_by_name('Sheet1')
# Inserire percorso repository
repo = git.Repo('/Users/patriziadecristofaro/GIT/RxJava')
logging.basicConfig(format='%(levelname)s %(asctime)s: %(message)s',level=logging.INFO, filename='debug.log')

print("+++++++++++++++++++++ INIZIO CALCOLO METRICHE +++++++++++++++++++++++++")
logging.info("+++++++++++++++++++++ INIZIO CALCOLO METRICHE +++++++++++++++++++++++++")
i = 293
c = ""
for x in range(296,976):
    try:
        if(i < 10): c = '00' + str(i)
        elif(i < 100 and i >= 10): c = '0' + str(i)
        else: c = str(i)
        comm=(sheet.cell(x,1).value)
        logging.info("processing "+ comm)
        a=repo.commit(comm)
        repo.git.checkout("-f",a)
        #repo.git.pull('origin', 'master')
        logging.info("Checkout succeded")
        subprocess.call(['java', '-jar',  'ck-0.6.5-SNAPSHOT-jar-with-dependencies.jar', '/Users/patriziadecristofaro/GIT/RxJava'])
        logging.info("Metrics calculated")
        shutil.move("class.csv", "commit_list/"+ c + "_" +comm+".csv")
        logging.info("File stored " + c + "_" + comm + ".csv")
        i += 1
    except(Exception) as e:
        logging.info(e)
        logging.info(traceback.format_exc())

print("+++++++++++++++++++++ FINE CALCOLO METRICHE +++++++++++++++++++++++++")
logging.info("+++++++++++++++++++++ FINE CALCOLO METRICHE +++++++++++++++++++++++++")


### CALCOLO RIGHE MODIFICATE CON GIT DIFF

print("+++++++++++++++++++++ INIZIO GIT DIFF +++++++++++++++++++++++")
logging.info("+++++++++++++++++++++ INIZIO GIT DIFF +++++++++++++++++++++++")

data = []
for x in range(1,976):
    try:
        comm1=(sheet.cell(x,1).value)
        comm2=(sheet.cell(x,2).value)
        logging.info("processing "+ comm1 + " and " + comm2)
        d = repo.git.diff("--stat", comm1, comm2)
        s = d.split("\n")
        s.pop()
        for l in s:
            ss = l.split('|')
            fname = "/RxJava/" + ss[0].strip()
            modLines = ss[1].strip().replace("+", "").replace("-", "")
            data1 = [comm1, fname, modLines]
            data.append(data1)
    except(Exception) as e:
        logging.info(e)
        logging.info(traceback.format_exc())
df = pd.DataFrame(data, columns = ['hash', 'file', 'modlines'])
df.to_csv("DiffDataset.csv", index = False)

print("+++++++++++++++++++++ FINE GIT DIFF +++++++++++++++++++++++")
logging.info("+++++++++++++++++++++ FINE GIT DIFF +++++++++++++++++++++++")



