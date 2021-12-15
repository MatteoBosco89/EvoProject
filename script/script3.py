import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from git import Repo
import git
import subprocess
import shutil
import logging
import traceback

file_excel = openpyxl.load_workbook('data.xlsx')
sheet = file_excel.get_sheet_by_name('Sheet1')
repo = git.Repo('/Users/patriziadecristofaro/Desktop/Arduino')
logging.basicConfig(format='%(levelname)s %(asctime)s: %(message)s',level=logging.INFO, filename='debug.log')

logging.info('Inizio')
for x in range(2,681):
    try:
        comm=(sheet.cell(x,1).value)
        logging.info("processing "+ comm)
        a=repo.commit(comm)
        repo.git.checkout(a)
        repo.git.pull('origin', 'master')
        logging.info("Pull succeded")
        subprocess.call(['java', '-jar', 'ck-0.6.5-SNAPSHOT-jar-with-dependencies.jar', '/Users/patriziadecristofaro/Desktop/Arduino'])
        logging.info("Metrics calculated")
        shutil.move("class.csv", "commit_list/"+comm+".csv")
        logging.info("File stored " + comm + ".csv")
    except(Exception) as e:
        logging.info(e)
        logging.info(traceback.format_exc())