import re
import openpyxl
from openpyxl import load_workbook
from pprint import pprint
from github import Github
import xlsxwriter

workbook = xlsxwriter.Workbook('data1.xlsx')
worksheet = workbook.add_worksheet()
file_excel = openpyxl.load_workbook('commit.xlsx')
sheet = file_excel.get_sheet_by_name('commit')
y=0
z=0
c=0
f=0
g=0
k=0
rem=dict()


for x in range(1,2046):
    a=(sheet.cell(x,4).value)
    b=(sheet.cell(x,1).value)
    

    try:
        s1=re.search(r"refact", a, re.I)
        s2=re.search(r"Fix", a, re.I)
        s3=re.search(r"bug", a, re.I)
        s4=re.search(r"issue", a, re.I)
        s5=re.search(r"remov", a, re.I)
        s6=re.search(r"resolv", a, re.I)
        
        
        if(s1 is not None):
            print(b)
            y=y+1
            rem[str(b)]=1
        if(s2 is not None):
            print(b)
            z=z+1
            rem[str(b)]=1
        if(s3 is not None):
            print(b)
            c=c+1
            rem[str(b)]=1
        if( s4 is not None):
            print(b)
            f=f+1
            rem[str(b)]=1
        if( s5 is not None):
            print(b)
            g=g+1
            rem[str(b)]=1
        if( s6 is not None):
            print(b)
            k=k+1
            rem[str(b)]=1

    except AttributeError:
        s1=re.search(r"Fixed",a,re.I)
        s2=re.search(r"Fix", a, re.I)
        s3=re.search(r"bug", a, re.I)
        s4=re.search(r"issue", a, re.I)
        s5=re.search(r"remov", a, re.I)
        s6=re.search(r"resolv", a, re.I)
        
print("regular expression with refactor = ",y)
print("regular expression with Fix = ",z)
print("regular expression with bug = ",c)
print("regular expression with issue = ",f)
print("regular expression with remov = ",g)
print("regular expression with resolv = ",k)
print("commit totali = ",len(rem))

row=0
col=0        
for key in rem.keys():
    row += 1
    worksheet.write(row, col, key)
    

workbook.close()